# analytics/views.py
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.db import connection
from django.http import HttpResponse
from datetime import datetime
import json
from .models import ReporteGenerado
from .serializers import ReporteGeneradoSerializer, ReporteSolicitudSerializer



class ReporteGeneradoListView(generics.ListAPIView):
    serializer_class = ReporteGeneradoSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return ReporteGenerado.objects.select_related('usuario').all()
        return ReporteGenerado.objects.filter(usuario=self.request.user).select_related('usuario')

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_reporte(request):
    serializer = ReporteSolicitudSerializer(data=request.data)
    if serializer.is_valid():
        datos = serializer.validated_data
        
        # Crear registro de reporte
        reporte = ReporteGenerado.objects.create(
            usuario=request.user,
            tipo_reporte=datos['tipo_reporte'],
            formato_salida=datos['formato_salida'],
            parametros=datos,
            estado='procesando'
        )
        
        # Generar consulta SQL basada en parámetros
        consulta_sql = construir_consulta_sql(datos)
        reporte.consulta_sql = consulta_sql
        
        try:
            # Ejecutar consulta y generar archivo
            if datos['formato_salida'] == 'pdf':
                archivo_url = generar_reporte_pdf(consulta_sql, reporte.id)
            elif datos['formato_salida'] == 'excel':
                archivo_url = generar_reporte_excel(consulta_sql, reporte.id)
            else:
                archivo_url = generar_reporte_csv(consulta_sql, reporte.id)
            
            reporte.url_descarga = archivo_url
            reporte.estado = 'completado'
            reporte.save()
            
            return Response(ReporteGeneradoSerializer(reporte).data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            reporte.estado = 'error'
            reporte.save()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

def construir_consulta_sql(parametros):
    tipo_reporte = parametros['tipo_reporte']
    fecha_inicio = parametros.get('fecha_inicio')
    fecha_fin = parametros.get('fecha_fin')
    
    consultas = {
        'ventas': """
            SELECT p.id, p.fecha_pedido, u.first_name, u.last_name, 
                    p.monto_total, p.estado_pedido
            FROM pedidos p
            JOIN usuarios u ON p.usuario_id = u.id
            WHERE 1=1
        """,
        'clientes': """
            SELECT u.id, u.first_name, u.last_name, u.email, u.telefono,
                    COUNT(p.id) as total_pedidos, SUM(p.monto_total) as total_gastado
            FROM usuarios u
            LEFT JOIN pedidos p ON u.id = p.usuario_id
            WHERE 1=1
            GROUP BY u.id, u.first_name, u.last_name, u.email, u.telefono
        """,
        'productos': """
            SELECT pr.id, pr.nombre, pr.sku, pr.precio, c.nombre_categoria,
                    i.stock_actual, COUNT(dp.id) as total_vendido
            FROM productos pr
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            LEFT JOIN inventario i ON pr.id = i.producto_id
            LEFT JOIN detalle_pedido dp ON pr.id = dp.producto_id
            WHERE 1=1
            GROUP BY pr.id, pr.nombre, pr.sku, pr.precio, c.nombre_categoria, i.stock_actual
        """
    }
    
    consulta_base = consultas.get(tipo_reporte, consultas['ventas'])
    
    # Aplicar filtros
    where_conditions = []
    if fecha_inicio:
        where_conditions.append(f"p.fecha_pedido >= '{fecha_inicio}'")
    if fecha_fin:
        where_conditions.append(f"p.fecha_pedido <= '{fecha_fin} 23:59:59'")
    
    if where_conditions:
        consulta_base += " AND " + " AND ".join(where_conditions)
    
    return consulta_base

def generar_reporte_pdf(consulta_sql, reporte_id):
    # Implementación básica - usar reportlab en producción
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from django.core.files.storage import default_storage
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    
    # Ejecutar consulta
    with connection.cursor() as cursor:
        cursor.execute(consulta_sql)
        resultados = cursor.fetchall()
        columnas = [col[0] for col in cursor.description]
    
    # Generar PDF básico
    p.drawString(100, 800, "Reporte SmartSales365")
    y = 780
    for fila in resultados[:20]:  # Limitar para ejemplo
        p.drawString(100, y, str(fila))
        y -= 20
    
    p.save()
    buffer.seek(0)
    
    # Guardar archivo
    nombre_archivo = f'reportes/reporte_{reporte_id}.pdf'
    default_storage.save(nombre_archivo, buffer)
    
    return default_storage.url(nombre_archivo)

def generar_reporte_excel(consulta_sql, reporte_id):
    # Implementación básica - usar openpyxl en producción
    import openpyxl
    from django.core.files.storage import default_storage
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Ejecutar consulta
    with connection.cursor() as cursor:
        cursor.execute(consulta_sql)
        resultados = cursor.fetchall()
        columnas = [col[0] for col in cursor.description]
    
    # Escribir encabezados
    for col_idx, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_idx, value=columna)
    
    # Escribir datos
    for row_idx, fila in enumerate(resultados, 2):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Guardar archivo
    nombre_archivo = f'reportes/reporte_{reporte_id}.xlsx'
    wb.save(nombre_archivo)
    
    return default_storage.url(nombre_archivo)

def generar_reporte_csv(consulta_sql, reporte_id):
    import csv
    from django.core.files.storage import default_storage
    
    # Ejecutar consulta
    with connection.cursor() as cursor:
        cursor.execute(consulta_sql)
        resultados = cursor.fetchall()
        columnas = [col[0] for col in cursor.description]
    
    # Generar CSV
    nombre_archivo = f'reportes/reporte_{reporte_id}.csv'
    with default_storage.open(nombre_archivo, 'w') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(columnas)
        writer.writerows(resultados)
    
    return default_storage.url(nombre_archivo)

# =============================================================================
# AGREGAR VIEWS DE REPORTES
# =============================================================================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_reporte_por_id(request, reporte_id):
    """Obtener reporte por ID"""
    reporte = ReporteGenerado.obtener_por_id(reporte_id)
    if reporte:
        # Verificar que el usuario tiene acceso al reporte
        if reporte.usuario == request.user or request.user.is_staff:
            serializer = ReporteGeneradoSerializer(reporte)
            return Response(serializer.data)
        else:
            return Response(
                {'error': 'No autorizado para ver este reporte'}, 
                status=status.HTTP_403_FORBIDDEN
            )
    else:
        return Response({'error': 'Reporte no encontrado'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def listar_reportes_usuario(request, usuario_id=None):
    """Listar reportes de usuario"""
    if usuario_id and request.user.is_staff:
        # Admin puede ver reportes de cualquier usuario
        target_user_id = usuario_id
    else:
        # Usuario normal solo puede ver sus propios reportes
        target_user_id = request.user.id
    
    reportes = ReporteGenerado.listar_por_usuario(target_user_id)
    serializer = ReporteGeneradoSerializer(reportes, many=True)
    return Response(serializer.data)